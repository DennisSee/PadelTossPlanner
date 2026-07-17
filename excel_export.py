"""Excel-export voor de TOS Padelplanner."""

from __future__ import annotations

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from planner import (
    Player,
    PlannerSettings,
    RoundPlan,
    player_statistics,
    schedule_rows,
)


DARK_BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
VERY_LIGHT_BLUE = "F3F7FA"
WHITE = "FFFFFF"
BORDER_COLOR = "D9E2F3"


def _style_header(ws, row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def _auto_fit(ws, max_width: int = 42) -> None:
    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        width = 10
        for cell in cells:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def build_excel_bytes(
    settings: PlannerSettings,
    players: Sequence[Player],
    courts: Sequence[str],
    rounds: Sequence[RoundPlan],
    diagnostics: dict[str, object],
) -> bytes:
    """Maak een volledig Excelbestand in het geheugen."""
    workbook = Workbook()
    schema_sheet = workbook.active
    schema_sheet.title = "Schema"
    schema_sheet.sheet_view.showGridLines = False
    schema_sheet.freeze_panes = "A6"

    schema_sheet["A1"] = "TOS padelschema"
    schema_sheet["A1"].font = Font(size=18, bold=True, color=DARK_BLUE)
    schema_sheet["A2"] = "Speelperiode"
    schema_sheet["B2"] = f"{settings.start_time:%H:%M} - {settings.end_time:%H:%M}"
    schema_sheet["D2"] = "Wedstrijdduur"
    schema_sheet["E2"] = f"{settings.match_minutes} minuten"
    schema_sheet["G2"] = "Rondes"
    schema_sheet["H2"] = diagnostics["rounds"]
    schema_sheet["A3"] = "Spelers"
    schema_sheet["B3"] = len(players)
    schema_sheet["D3"] = "Banen"
    schema_sheet["E3"] = len(courts)
    schema_sheet["G3"] = "Onbenutte tijd"
    schema_sheet["H3"] = f"{diagnostics['unused_minutes']} minuten"

    rows = schedule_rows(rounds, courts, players, settings)
    headers = list(rows[0].keys())
    for column, header in enumerate(headers, start=1):
        schema_sheet.cell(5, column, header)
    _style_header(schema_sheet, 5, 1, len(headers))

    for row_index, row in enumerate(rows, start=6):
        for column, header in enumerate(headers, start=1):
            schema_sheet.cell(row_index, column, row[header])
        if row_index % 2 == 0:
            for column in range(1, len(headers) + 1):
                schema_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor=VERY_LIGHT_BLUE
                )
        for column in range(1, len(headers) + 1):
            schema_sheet.cell(row_index, column).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    schema_sheet.auto_filter.ref = f"A5:I{5 + len(rows)}"
    widths = {
        "A": 9,
        "B": 17,
        "C": 23,
        "D": 28,
        "E": 12,
        "F": 28,
        "G": 12,
        "H": 14,
        "I": 30,
    }
    for letter, width in widths.items():
        schema_sheet.column_dimensions[letter].width = width

    stats_sheet = workbook.create_sheet("Spelerstatistiek")
    stats_sheet.sheet_view.showGridLines = False
    stats_sheet.freeze_panes = "A3"
    stats_sheet["A1"] = "Spelerstatistiek"
    stats_sheet["A1"].font = Font(size=18, bold=True, color=DARK_BLUE)

    stats = player_statistics(rounds, players, diagnostics)
    stat_headers = list(stats[0].keys())
    for column, header in enumerate(stat_headers, start=1):
        stats_sheet.cell(2, column, header)
    _style_header(stats_sheet, 2, 1, len(stat_headers))

    for row_index, row in enumerate(stats, start=3):
        for column, header in enumerate(stat_headers, start=1):
            stats_sheet.cell(row_index, column, row[header])
        if row_index % 2 == 1:
            for column in range(1, len(stat_headers) + 1):
                stats_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor=VERY_LIGHT_BLUE
                )
        for column in range(1, len(stat_headers) + 1):
            stats_sheet.cell(row_index, column).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    stats_sheet.auto_filter.ref = f"A2:G{2 + len(stats)}"
    _auto_fit(stats_sheet)
    stats_sheet.column_dimensions["G"].width = 46

    input_sheet = workbook.create_sheet("Invoer")
    input_sheet.sheet_view.showGridLines = False
    input_sheet["A1"] = "Gebruikte invoer"
    input_sheet["A1"].font = Font(size=18, bold=True, color=DARK_BLUE)
    input_sheet["A3"] = "Instelling"
    input_sheet["B3"] = "Waarde"
    _style_header(input_sheet, 3, 1, 2)
    input_rows = [
        ("Starttijd", settings.start_time.strftime("%H:%M")),
        ("Eindtijd", settings.end_time.strftime("%H:%M")),
        ("Wedstrijdduur", settings.match_minutes),
        ("Dubbele partners toegestaan", "Ja" if settings.allow_repeat_partners else "Nee"),
        ("Banen", ", ".join(courts)),
    ]
    for row_index, values in enumerate(input_rows, start=4):
        input_sheet.cell(row_index, 1, values[0])
        input_sheet.cell(row_index, 2, values[1])
        input_sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        input_sheet.cell(row_index, 1).font = Font(bold=True)

    player_start = 11
    input_sheet.cell(player_start, 1, "Speler")
    input_sheet.cell(player_start, 2, "Ranking")
    _style_header(input_sheet, player_start, 1, 2)
    for row_index, player in enumerate(players, start=player_start + 1):
        input_sheet.cell(row_index, 1, player.name)
        input_sheet.cell(row_index, 2, player.ranking)
    _auto_fit(input_sheet)
    input_sheet.column_dimensions["A"].width = 30
    input_sheet.column_dimensions["B"].width = 55

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
